import openpyxl as xl

file1 = input("Enter file 1: ")
file2 = input("Enter file 2: ")

filename = file1
workBook1 = xl.load_workbook(filename)
workSheet1 = workBook1.worksheets[0]

filename = file2
workBook2 = xl.load_workbook(filename)
workSheet2 = workBook2.worksheets[0]

maxRows = workSheet1.max_row
maxColumns = workSheet1.max_column

print("total rows and columsn in first file", maxRows, maxColumns)

sheetOneValue = []

for i in range(1, maxRows + 1):
    currentLine = ""
    for j in range(1, maxColumns + 1):
        currentLine = currentLine + str(workSheet1.cell(row=i, column=j).value) + " "
    sheetOneValue.append(currentLine)

print("completed reading first file")

sheetTwoValue = []
maxRows = workSheet2.max_row
maxColumns = workSheet2.max_column

print("totla rows and columns in second file", maxRows, maxColumns)

for i in range(1, maxRows + 1):
    currentLine = ""
    for j in range(1, maxColumns + 1):
        currentLine = currentLine + str(workSheet2.cell(row=i, column=j).value) + " "
    sheetTwoValue.append(currentLine)

print("completed reading second file")

print("File 1 has " + str(len(sheetOneValue)) + " lines")
print("File 2 has " + str(len(sheetTwoValue)) + " lines")

print("comparing two files")

count = 0
sheetOneValue = list(sheetOneValue)
sheetTwoValue = list(sheetTwoValue)


for line in sheetOneValue:
    if line not in sheetTwoValue:
        #print(line)
        count+=1

print("total lines not in file 2: " + str(count))